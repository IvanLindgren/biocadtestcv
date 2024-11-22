# report_generation.py
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd
from utils import OUTPUT_DIR, REFERENCE_CLARINASE, REFERENCE_CLARITINE


def create_similarity_chart(folder: str, results: list[tuple[str, float]], output_dir: Path):
    """Создание столбчатой диаграммы схожести."""
    filenames, scores = zip(*results)
    plt.figure(figsize=(10, 6))
    plt.bar(filenames, scores, color='skyblue')
    plt.xlabel("Изображения")
    plt.ylabel("Схожесть (%)")
    plt.title(f"Схожесть изображений в папке {folder} с референсным")
    plt.xticks(rotation=90, fontsize=8)
    plt.tight_layout()
    chart_path = output_dir / f"similarity_chart_{folder.replace(' ', '_')}.png"
    plt.savefig(chart_path)
    plt.close()
    return chart_path

def create_excel_report(results: dict, output_dir: Path):
    """Создание отчета в формате Excel."""
    wb = openpyxl.Workbook()

    for folder, folder_results in results.items():
        ws = wb.create_sheet(title=folder[:31])  # Ограничение длины имени листа
        ws.append(["Имя файла", "Схожесть (%)", "Различия"])

        # Установка ширины столбцов
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20

        # Выравнивание заголовков по центру
        for col in range(1, 4):
             ws[get_column_letter(col) + '1'].alignment = Alignment(horizontal='center')

        row_num = 2
        for filename, score in folder_results:
            ws.cell(row=row_num, column=1, value=filename)
            ws.cell(row=row_num, column=2, value=f"{score:.2f}%")
            diff_image_path = output_dir / f"diff_{folder}_{filename.replace(' ', '_').split('.')[0]}.png"
            if diff_image_path.exists():
                img = ExcelImage(str(diff_image_path))
                img.width = 100
                img.height = 100
                ws.add_image(img, f'C{row_num}')
                ws.row_dimensions[row_num].height = 80  # Увеличиваем высоту строки для изображения
            row_num += 1

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']  # Удаляем лист по умолчанию

    excel_path = output_dir / "similarity_report.xlsx"
    wb.save(excel_path)
    return excel_path

def generate_text_report(results: dict, output_dir: Path, excel_path: Path, charts_paths: dict):
    """Генерация текстового отчета о результатах сравнения."""
    with open(output_dir / "report.txt", "w", encoding='utf-8') as f:
        f.write("Отчет о сравнении изображений упаковок\n\n")
        f.write("Цель задания:\n")
        f.write(
            "   Сравнить изображения упаковок из папок “Clarinase 14 repetabs” и ”Claritine 20 tablets”, определив степень их схожести и выделив различия.\n\n")
        f.write("Описание задачи:\n")
        f.write(
            "   Вам необходимо обработать изображения из двух папок: Clarinaze14 и Claritine20.\n   Основная задача состоит в следующем:\n")
        f.write("   1. Сопоставление упаковок:\n")
        f.write(
            f"      Выберите пачку на изображении из каждой папки: Clarinase 14 repetabs: “{REFERENCE_CLARINASE}” и Claritine 20 tablets: \"{REFERENCE_CLARITINE}\"\n")
        f.write("   2. Преобразование изображений:\n")
        f.write("      Преобразуйте остальные изображения в каждой папке к единому формату.\n")
        f.write("   3. Определение степени схожести:\n")
        f.write(
            "      Для выбранного изображения из папки Clarinaze14 сравните его с оставшимися изображениями в обеих папках. Используйте алгоритмы сравнения изображений для определения степени схожести, выраженной в баллах (от 0 до 100%).\n")
        f.write("   4. Выделение различий:\n")
        f.write(
            "      Если имеются различия между сравниваемыми изображениями, выявите их и выделите визуально на конечном изображении или в отчете.\n\n")
        f.write("Методы анализа:\n")
        f.write(
            "   Для сравнения изображений использовался алгоритм структурного подобия (SSIM). Изображения предварительно преобразовывались в оттенки серого и изменялся их размер до 500x500 пикселей. SSIM вычисляется для каждого изображения в сравнении с референсным изображением. Значение SSIM, выраженное в процентах, показывает степень схожести изображений (от 0 до 100%). Различия между изображениями визуализируются, если SSIM ниже 90%, с выделением области различий красной рамкой.\n\n")
        f.write("Результаты сравнения:\n")

        for folder, folder_results in results.items():
            f.write(f"Результаты для папки: {folder}\n")
            f.write("-----------------------------------------\n")
            f.write("{:<40} {:<15}\n".format("Имя файла", "Схожесть (%)"))
            f.write("-----------------------------------------\n")

            for filename, score in folder_results:
                f.write("{:<40} {:<15.2f}\n".format(filename, score))
            f.write("\n")
            f.write(f"График схожести для папки {folder}: {charts_paths[folder]}\n")
        f.write(
            f"Визуализация различий для изображений с низкой схожестью (менее 90%) сохранена в папке {OUTPUT_DIR}.\n")
        f.write(f"Excel отчет с результатами и визуализацией: {excel_path}\n")
        f.write("Файлы с результатами схожести сохранены в формате .npy в папке results.\n")

    print("Отчет сохранен в файл report.txt")
